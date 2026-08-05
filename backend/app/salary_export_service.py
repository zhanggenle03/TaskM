"""薪资记录导出服务 —— 生成 XLSX 格式的多 Sheet 导出报告

Sheet 结构：
  1. 总览 —— 区间汇总 + 个税年度汇算
  2. 明细 —— 各月薪资明细（扁平化表格）
  3. 五险一金变化 —— 各险种基数与比例逐月变化追踪
"""

import io
import os
from datetime import datetime
from typing import List, Optional, Dict
from decimal import Decimal, ROUND_HALF_UP

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

from sqlalchemy.orm import Session

from .database import SalaryRecord

# ── 五险一金标准键名 —— 与 salary.py 保持一致 ──
SOCIAL_RATE_KEYS = [
    "养老保险(个人)", "医疗保险(个人)", "失业保险(个人)", "住房公积金(个人)",
    "养老保险(公司)", "医疗保险(公司)", "失业保险(公司)",
    "工伤保险(公司)", "生育保险(公司)", "住房公积金(公司)",
]

# ── 工资条图片导出参数 ──
SLIP_COL_WIDTH = 34   # 工资条列宽（字符数）
SLIP_MAX_W = 220      # 图片最大宽度（px，等比缩放不放大）
SLIP_MAX_H = 170      # 图片最大高度（px，等比缩放不放大）

# ── 字体与颜色常量 ──
FONT_TITLE = Font(name='微软雅黑', size=16, bold=True, color='1F3864')
FONT_SECTION = Font(name='微软雅黑', size=13, bold=True, color='2F5496')
FONT_HEADER = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
FONT_BODY = Font(name='微软雅黑', size=10)
FONT_BODY_BOLD = Font(name='微软雅黑', size=10, bold=True)
FONT_MUTED = Font(name='微软雅黑', size=10, color='999999')

FILL_HEADER = PatternFill('solid', fgColor='4472C4')
FILL_SUBHEADER = PatternFill('solid', fgColor='D6E4F0')
FILL_LIGHT = PatternFill('solid', fgColor='F2F2F2')
FILL_INCOME = PatternFill('solid', fgColor='E8F5E9')      # 收入浅绿
FILL_DEDUCTION = PatternFill('solid', fgColor='FFF3E0')   # 扣款浅橙
FILL_TAX = PatternFill('solid', fgColor='FFEBEE')         # 个税浅红
FILL_COMPANY = PatternFill('solid', fgColor='ECEFF1')     # 公司承担浅灰

CATEGORY_FILLS = {
    'income': FILL_INCOME,
    'deduction': FILL_DEDUCTION,
    'tax': FILL_TAX,
    'company_cost': FILL_COMPANY,
}
CATEGORY_LABELS = {
    'income': '收入', 'deduction': '个人扣款', 'tax': '个税', 'company_cost': '公司承担',
}

THIN_BORDER = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0'),
)

ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')


def _fmt(n) -> str:
    return f"¥{n:,.2f}" if n is not None else "—"


def _apply_cell(ws, row, col, value, font=FONT_BODY, fill=None,
                alignment=ALIGN_CENTER, border=THIN_BORDER, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = alignment
    cell.border = border
    if number_format:
        cell.number_format = number_format
    return cell


def _load_slip_image(file_path: str) -> Image:
    """加载工资条图片；openpyxl 不支持 webp 等格式时经 Pillow 转 PNG 返回。"""
    from PIL import Image as PILImage
    pil = PILImage.open(file_path)
    pil.load()
    if (pil.format or '').upper() in ('PNG', 'JPEG', 'GIF'):
        pil.close()
        return Image(file_path)
    buf = io.BytesIO()
    mode = pil.mode if pil.mode in ('RGBA', 'LA') else 'RGB'
    pil.convert(mode).save(buf, format='PNG')
    buf.seek(0)
    return Image(buf)


def _write_header_row(ws, row, headers, start_col=1, fill=FILL_HEADER, font=FONT_HEADER):
    for i, h in enumerate(headers):
        _apply_cell(ws, row, start_col + i, h, font=font, fill=fill, alignment=ALIGN_CENTER)


def _write_section_title(ws, row, title, col=1):
    _apply_cell(ws, row, col, title, font=FONT_SECTION, fill=None,
                alignment=Alignment(horizontal='left', vertical='center'),
                border=Border())
    return row + 1


def _compute_totals(record):
    gross = sum(i.amount for i in record.items if i.category == "income")
    personal_deduction = sum(i.amount for i in record.items if i.category in ("deduction", "tax"))
    company_cost = sum(i.amount for i in record.items if i.category == "company_cost")
    net = gross - personal_deduction
    return gross, personal_deduction, net, company_cost


def _build_summary_sheet(ws, summary: dict, tax_summary: Optional[dict],
                          period_from: Optional[str], period_to: Optional[str]):
    """填充 Sheet 1：总览"""
    # 标题
    _apply_cell(ws, 1, 1, '薪资记录导出报告', font=FONT_TITLE,
                alignment=Alignment(horizontal='left', vertical='center'), border=Border())
    _apply_cell(ws, 2, 1, f'数据区间：{period_from or "最早"} ~ {period_to or "最晚"}    '
                f'导出时间：{datetime.now().strftime("%Y-%m-%d %H:%M")}',
                font=FONT_MUTED,
                alignment=Alignment(horizontal='left', vertical='center'), border=Border())
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)

    # ── 区间汇总 ——
    r = 4
    _apply_cell(ws, r, 1, '区间汇总', font=FONT_SECTION, border=Border())
    r += 1

    sum_headers = ['指标', '金额（元）']
    _write_header_row(ws, r, sum_headers)
    r += 1

    sum_rows = [
        ('记录月份数', f'{summary.get("record_count", 0)} 个月'),
        ('区间应发合计', summary.get('total_gross', 0)),
        ('应扣合计', summary.get('total_personal_deduction', 0)),
        ('区间实发合计', summary.get('total_net', 0)),
        ('到账合计', summary.get('total_credited', 0)),
        ('公司承担合计', summary.get('total_company_cost', 0)),
        ('月均实发', summary.get('avg_net', 0)),
        ('实际个税合计', summary.get('total_actual_tax', 0)),
    ]
    for i, (label, val) in enumerate(sum_rows):
        bg = FILL_LIGHT if i % 2 == 1 else None
        _apply_cell(ws, r, 1, label, font=FONT_BODY_BOLD, fill=bg, alignment=ALIGN_LEFT)
        if isinstance(val, str):
            _apply_cell(ws, r, 2, val, font=FONT_BODY, fill=bg)
        else:
            _apply_cell(ws, r, 2, val, font=FONT_BODY, fill=bg,
                        number_format='#,##0.00')
        r += 1

    # ── 个税年度汇算 ——
    if tax_summary and tax_summary.get('month_count', 0) > 0:
        r += 1
        _apply_cell(ws, r, 1, '个税年度汇算', font=FONT_SECTION, border=Border())
        r += 1

        tax_headers = ['指标', '数值']
        _write_header_row(ws, r, tax_headers)
        r += 1

        tax_rows = [
            ('年度', str(tax_summary.get('year', ''))),
            ('有记录月份数', f'{tax_summary.get("month_count", 0)} 个月'),
            ('年度累计应发', tax_summary.get('total_gross', 0)),
            ('累计专项扣除', tax_summary.get('total_social_insurance', 0)),
            ('应纳税所得额', tax_summary.get('taxable_income', 0)),
            ('适用税率', tax_summary.get('tax_rate_label', '0%')),
            ('速算扣除数', tax_summary.get('quick_deduction', 0)),
        ]
        for i, (label, val) in enumerate(tax_rows):
            bg = FILL_LIGHT if i % 2 == 1 else None
            _apply_cell(ws, r, 1, label, font=FONT_BODY_BOLD, fill=bg, alignment=ALIGN_LEFT)
            if isinstance(val, str):
                _apply_cell(ws, r, 2, val, font=FONT_BODY, fill=bg)
            else:
                _apply_cell(ws, r, 2, val, font=FONT_BODY, fill=bg,
                            number_format='#,##0.00')
            r += 1

    # 列宽
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 20


def _build_detail_sheet(ws, records: List[SalaryRecord]):
    """填充 Sheet 2：明细（宽表格式，每月份一行，各项目为独立列）"""
    sorted_records = sorted(records, key=lambda x: x.period)

    # 类别标签与颜色映射
    CAT_LABELS = {
        'income': '收入', 'deduction': '个人扣款', 'tax': '个税', 'company_cost': '公司承担',
    }
    CAT_HEADER_FILLS = {
        'income': 'C8E6C9',    # 收入浅绿
        'deduction': 'FFE0B2', # 扣款浅橙
        'tax': 'FFCDD2',       # 个税浅红
        'company_cost': 'CFD8DC',  # 公司浅灰
    }

    # 收集所有出现过的项目名称及对应类别（按 income→deduction→tax→company_cost 排序）。
    # 同名收入项按计税状态拆分（计税/非计税各一列），避免混用时错位。
    cat_order = {'income': 0, 'deduction': 1, 'tax': 2, 'company_cost': 3}
    item_info = {}  # (name, tax_flag) -> {name, category, cat_label, taxable}
    for rec in sorted_records:
        for it in sorted(rec.items, key=lambda x: (cat_order.get(x.category, 9), x.sort_order)):
            if not it.name:
                continue
            tax_flag = 'nontax' if (it.category == 'income' and it.taxable is False) else 'tax'
            key = (it.name, tax_flag)
            if key not in item_info:
                item_info[key] = {
                    'name': it.name,
                    'category': it.category,
                    'cat_label': CAT_LABELS.get(it.category, it.category),
                    'taxable': tax_flag == 'nontax',
                }

    # 按类别+首次出现排序
    item_order = sorted(item_info.keys(), key=lambda k: (
        cat_order.get(item_info[k]['category'], 9),
        list(item_info.keys()).index(k),
    ))

    # 表头：基本信息列 + "类别-项目名"形式列（非计税收入带前缀） + 汇总列 + 工资条列
    base_headers = ['月份', '发放日期', '单位']
    item_headers = []
    for k in item_order:
        info = item_info[k]
        if info['category'] == 'income' and info['taxable']:
            item_headers.append(f"收入（非计税）-{info['name']}")
        else:
            item_headers.append(f"{info['cat_label']}-{info['name']}")
    sum_headers = ['应发合计', '应扣合计', '实发', '公司承担', '到账', '实际个税', '备注']
    # 工资条列：列数 = 单月最大附件数；某月多张按序占多列（第 i 张放第 i 列），不是每月都新开列
    max_slips = max((len(rec.slips) for rec in sorted_records), default=0)
    slip_headers = [f'工资条{i + 1}' for i in range(max_slips)]
    headers = base_headers + item_headers + sum_headers + slip_headers
    n_base = len(base_headers)
    n_item = len(item_headers)
    n_sum = len(sum_headers)
    slip_col_start = n_base + n_item + n_sum + 1  # 1-based：工资条首列

    # 写入表头
    _write_header_row(ws, 1, headers)
    ws.row_dimensions[1].height = 28

    # 表头分三层着色
    for ci, h in enumerate(headers):
        cell = ws.cell(row=1, column=ci + 1)
        if ci < n_base:
            cell.fill = PatternFill('solid', fgColor='B4C6E7')
        elif ci < n_base + n_item:
            key = item_order[ci - n_base]
            cat = item_info[key]['category']
            cell.fill = PatternFill('solid', fgColor=CAT_HEADER_FILLS.get(cat, 'B0BEC5'))
        else:
            cell.fill = PatternFill('solid', fgColor='7B9CD6')
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

    # 写入数据
    r = 2
    for rec in sorted_records:
        gross, pd, net, cc = _compute_totals(rec)

        # 构建 (name, tax_flag) → 金额映射（同名同计税状态累加）
        item_amt = {}
        tax_from_items = 0.0
        for it in rec.items:
            if it.name:
                tax_flag = 'nontax' if (it.category == 'income' and it.taxable is False) else 'tax'
                key = (it.name, tax_flag)
                item_amt[key] = item_amt.get(key, 0.0) + (it.amount or 0.0)
            if it.category == 'tax':
                tax_from_items += it.amount or 0.0

        # 如果记录没有单独的"个税"项行，但 actual_tax 有值，则补充
        if ('个税', 'tax') not in item_amt and rec.actual_tax is not None:
            item_amt[('个税', 'tax')] = rec.actual_tax

        # 基本信息列
        _apply_cell(ws, r, 1, rec.period, font=FONT_BODY)
        _apply_cell(ws, r, 2, rec.pay_date.strftime('%Y-%m-%d') if rec.pay_date else '', font=FONT_BODY)
        _apply_cell(ws, r, 3, rec.employer or '—', font=FONT_BODY, alignment=ALIGN_LEFT)

        # 项目金额列（按 item_order 顺序输出）
        for ci, key in enumerate(item_order):
            col = n_base + ci + 1
            amt = item_amt.get(key)
            if amt is not None:
                _apply_cell(ws, r, col, amt, font=FONT_BODY, number_format='#,##0.00')
            else:
                _apply_cell(ws, r, col, '', font=FONT_BODY)

        # 汇总列
        sum_col_start = n_base + n_item + 1
        sum_values = [
            gross, pd, net, cc,
            rec.credited_amount if rec.credited_amount is not None else '—',
            rec.actual_tax if rec.actual_tax is not None else '—',
            rec.remark or '',
        ]
        for ci, val in enumerate(sum_values):
            col = sum_col_start + ci
            is_num = isinstance(val, (int, float))
            _apply_cell(ws, r, col, val, font=FONT_BODY_BOLD,
                        number_format='#,##0.00' if is_num else None,
                        alignment=ALIGN_LEFT if not is_num else ALIGN_CENTER)

        # 工资条：第 i 张附件插到第 i 列（缺失/损坏文件跳过）
        if rec.slips:
            max_h = 0
            for i, slip in enumerate(sorted(rec.slips, key=lambda x: x.id)):
                if i >= max_slips or not slip.file_path or not os.path.exists(slip.file_path):
                    continue
                try:
                    img = _load_slip_image(slip.file_path)
                    ratio = min(SLIP_MAX_W / img.width, SLIP_MAX_H / img.height, 1.0)
                    img.width = max(1, int(img.width * ratio))
                    img.height = max(1, int(img.height * ratio))
                    ws.add_image(img, f"{get_column_letter(slip_col_start + i)}{r}")
                    max_h = max(max_h, img.height)
                except Exception:
                    continue
            if max_h:
                ws.row_dimensions[r].height = max_h + 8

        # 隔行底色
        if (r - 2) % 2 == 1:
            for ci in range(1, len(headers) + 1):
                ws.cell(row=r, column=ci).fill = FILL_LIGHT

        r += 1

    # 列宽
    for ci in range(len(headers)):
        col_letter = get_column_letter(ci + 1)
        if ci < n_base:
            widths = [10, 12, 16]
            ws.column_dimensions[col_letter].width = widths[ci]
        elif ci < n_base + n_item:
            ws.column_dimensions[col_letter].width = 18  # 项目列（含前缀）
        elif ci < n_base + n_item + n_sum:
            sum_widths = [12, 12, 12, 12, 12, 12, 20]
            ws.column_dimensions[col_letter].width = sum_widths[ci - (n_base + n_item)]
        else:
            ws.column_dimensions[col_letter].width = SLIP_COL_WIDTH  # 工资条图片列

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 自动筛选
    max_row = r - 1
    if max_row > 1:
        last_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f'A1:{last_col}{max_row}'


def _build_social_sheet(ws, records: List[SalaryRecord]):
    """填充 Sheet 3：五险一金基数与比例变化（按险种分组，个人与公司横向对比）"""
    # 收集数据
    changes = {}
    for rec in sorted(records, key=lambda r: r.period):
        for item in sorted(rec.items, key=lambda x: x.sort_order):
            if item.name in SOCIAL_RATE_KEYS:
                if item.name not in changes:
                    changes[item.name] = []
                changes[item.name].append({
                    'period': rec.period,
                    'base': item.base,
                    'rate': item.rate,
                    'amount': item.amount,
                })

    if not changes:
        _apply_cell(ws, 1, 1, '所选区间内无五险一金相关记录。', font=FONT_MUTED,
                    alignment=Alignment(horizontal='left', vertical='center'), border=Border())
        return

    FILL_CHANGED = PatternFill('solid', fgColor='FFF9C4')
    FILL_FIRST = PatternFill('solid', fgColor='E3F2FD')

    # 险种配对：(显示名, 个人项目名, 公司项目名)
    SOCIAL_PAIRS = [
        ('养老', '养老保险(个人)', '养老保险(公司)'),
        ('医疗', '医疗保险(个人)', '医疗保险(公司)'),
        ('失业', '失业保险(个人)', '失业保险(公司)'),
        ('工伤', None, '工伤保险(公司)'),
        ('生育', None, '生育保险(公司)'),
        ('公积金', '住房公积金(个人)', '住房公积金(公司)'),
    ]

    # 标题
    _apply_cell(ws, 1, 1, '五险一金基数与比例变化跟踪', font=FONT_TITLE,
                alignment=Alignment(horizontal='left', vertical='center'), border=Border())
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)

    # 列头
    col_headers = ['月份',
                   '个人基数', '个人比例%', '个人金额',
                   '公司基数', '公司比例%', '公司金额', '合计']
    n_data_cols = len(col_headers)

    r = 3

    def _find(entries, period):
        for e in entries:
            if e['period'] == period:
                return e
        return None

    all_periods = sorted(set(e['period'] for entries in changes.values() for e in entries))

    for pair_name, personal_name, company_name in SOCIAL_PAIRS:
        # 收集该险种所有月份的数据
        per_entries = changes.get(personal_name, []) if personal_name else []
        com_entries = changes.get(company_name, []) if company_name else []
        if not per_entries and not com_entries:
            continue

        # 合并所有月份（取个人或公司有的月份）
        pair_periods = sorted(set(e['period'] for e in per_entries + com_entries))

        # 险种标题行
        _apply_cell(ws, r, 1, f'▎{pair_name}', font=FONT_SECTION, border=Border())
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_data_cols)
        r += 1

        # 列头
        _write_header_row(ws, r, col_headers)
        r += 1

        for pi, period in enumerate(pair_periods):
            per = _find(per_entries, period) if personal_name else None
            com = _find(com_entries, period) if company_name else None

            prev_period = pair_periods[pi - 1] if pi > 0 else None
            prev_per = _find(per_entries, prev_period) if (personal_name and prev_period) else None
            prev_com = _find(com_entries, prev_period) if (company_name and prev_period) else None

            is_first = (pi == 0)
            row_bg = FILL_FIRST if is_first else (FILL_LIGHT if r % 2 == 0 else None)

            # 月份
            _apply_cell(ws, r, 1, period, font=FONT_BODY_BOLD, fill=row_bg)

            # 个人数据
            p_base = per['base'] if per else None
            p_rate = per['rate'] if per else None
            p_amt = per['amount'] if per else None

            p_base_chg = per and prev_per and prev_per.get('base') is not None and per['base'] is not None and abs(per['base'] - prev_per['base']) > 0.0001
            p_rate_chg = per and prev_per and prev_per.get('rate') is not None and per['rate'] is not None and abs(per['rate'] - prev_per['rate']) > 0.0001
            p_amt_chg = per and prev_per and abs(per['amount'] - prev_per['amount']) > 0.0001

            for ci, (val, chg, prev_v, nf) in enumerate([
                (p_base, p_base_chg, prev_per['base'] if prev_per else None, '#,##0.00'),
                (p_rate, p_rate_chg, prev_per['rate'] if prev_per else None, '0.00'),
                (p_amt, p_amt_chg, prev_per['amount'] if prev_per else None, '#,##0.00'),
            ], start=2):
                if chg and val is not None:
                    diff = round(val - prev_v, 4)
                    d = f'+{abs(diff):,.2f}' if diff > 0 else f'-{abs(diff):,.2f}'
                    cell = ws.cell(row=r, column=ci)
                    cell.value = f'{val}  ({d})'
                    cell.font = Font(name='微软雅黑', size=10, bold=True,
                                     color='C62828' if diff > 0 else '2E7D32')
                    cell.fill = FILL_CHANGED
                    cell.border = THIN_BORDER
                    cell.alignment = ALIGN_CENTER
                else:
                    _apply_cell(ws, r, ci, val if val is not None else '—',
                                font=FONT_BODY, fill=row_bg,
                                number_format=nf if val is not None else None)

            # 公司数据
            c_base = com['base'] if com else None
            c_rate = com['rate'] if com else None
            c_amt = com['amount'] if com else None

            c_base_chg = com and prev_com and prev_com.get('base') is not None and com['base'] is not None and abs(com['base'] - prev_com['base']) > 0.0001
            c_rate_chg = com and prev_com and prev_com.get('rate') is not None and com['rate'] is not None and abs(com['rate'] - prev_com['rate']) > 0.0001
            c_amt_chg = com and prev_com and abs(com['amount'] - prev_com['amount']) > 0.0001

            for ci, (val, chg, prev_v, nf) in enumerate([
                (c_base, c_base_chg, prev_com['base'] if prev_com else None, '#,##0.00'),
                (c_rate, c_rate_chg, prev_com['rate'] if prev_com else None, '0.00'),
                (c_amt, c_amt_chg, prev_com['amount'] if prev_com else None, '#,##0.00'),
            ], start=5):
                if chg and val is not None:
                    diff = round(val - prev_v, 4)
                    d = f'+{abs(diff):,.2f}' if diff > 0 else f'-{abs(diff):,.2f}'
                    cell = ws.cell(row=r, column=ci)
                    cell.value = f'{val}  ({d})'
                    cell.font = Font(name='微软雅黑', size=10, bold=True,
                                     color='C62828' if diff > 0 else '2E7D32')
                    cell.fill = FILL_CHANGED
                    cell.border = THIN_BORDER
                    cell.alignment = ALIGN_CENTER
                else:
                    _apply_cell(ws, r, ci, val if val is not None else '—',
                                font=FONT_BODY, fill=row_bg,
                                number_format=nf if val is not None else None)

            # 合计
            total = round((p_amt or 0) + (c_amt or 0), 2)
            _apply_cell(ws, r, 8, total, font=FONT_BODY_BOLD, fill=row_bg,
                        number_format='#,##0.00')

            r += 1

        r += 1  # 险种之间空行

    # 列宽
    widths_list = [10, 16, 10, 14, 16, 10, 14, 14]
    for i, w in enumerate(widths_list):
        ws.column_dimensions[get_column_letter(i + 1)].width = w


def generate_salary_export(
    db: Session,
    records: List[SalaryRecord],
    period_from: Optional[str],
    period_to: Optional[str],
    summary: dict,
    tax_summary: Optional[dict],
    salary_config: Optional[dict],
) -> bytes:
    """生成薪资导出 XLSX 多 Sheet 工作簿，返回字节内容。"""
    wb = Workbook()

    # ── Sheet 1: 总览 ──
    ws1 = wb.active
    ws1.title = '总览'
    _build_summary_sheet(ws1, summary, tax_summary, period_from, period_to)

    # ── Sheet 2: 明细 ──
    ws2 = wb.create_sheet('明细')
    _build_detail_sheet(ws2, records)

    # ── Sheet 3: 五险一金变化 ──
    ws3 = wb.create_sheet('五险一金变化')
    _build_social_sheet(ws3, records)

    # ── 输出字节 ──
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
