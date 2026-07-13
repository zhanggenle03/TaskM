"""出勤明细 Excel 导出（后端 openpyxl 生成，含原生 Excel 图表）

前端把已算好的计算结果 POST 过来，后端用 openpyxl 渲染成多工作表 xlsx：
- 出勤明细（逐日，真实日期列 + 按自然周浅蓝间隔底色）
- 总统计 / 按月统计 / 按项目统计
- 统计图表：原生 BarChart（按月趋势）+ 原生 PieChart（类型分布），Excel/WPS 打开即真实可交互图表
"""
import io
import urllib.parse
from datetime import date, datetime, timedelta
from typing import Any, Dict, Optional

from fastapi import APIRouter, HTTPException
from fastapi.responses import Response
import openpyxl
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/attendance", tags=["attendance-export"])

WEEKDAY_CN = ["一", "二", "三", "四", "五", "六", "日"]  # Python weekday(): 周一=0
# 浅蓝填充（表头与周间隔带共用）
_BLUE = PatternFill(start_color="FFDCE6F1", end_color="FFDCE6F1", fill_type="solid")


def _parse_date(s: str) -> Optional[date]:
    try:
        y, m, d = map(int, str(s).split("-"))
        return date(y, m, d)
    except Exception:
        return None


def _header_style(ws, ncols: int, nrows: int):
    """表头：加粗+居中+浅蓝填充（仅覆盖有内容的单元格）；冻结首行 + 自动筛选。"""
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = _BLUE
    ws.freeze_panes = "A2"
    if nrows:
        ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{nrows + 1}"


@router.post("/export-excel")
async def export_attendance_excel(payload: Dict[str, Any]):
    try:
        start = payload.get("start") or ""
        end = payload.get("end") or ""
        result = payload.get("result") or {}
        days = result.get("days") or []
        total = result.get("total") or {}
        monthly = result.get("monthly") or []
        by_project = result.get("byProject") or []

        wb = openpyxl.Workbook()

        # 1) 出勤明细
        ws = wb.active
        ws.title = "出勤明细"
        headers = ["日期", "星期", "类型", "涉及项目", "人天", "人天说明", "工作记录", "是否预估"]
        ws.append(headers)
        _WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
        first_monday = None
        for i, d in enumerate(days):
            dt = _parse_date(d.get("date", ""))
            if dt is None:
                continue
            if first_monday is None:
                first_monday = dt - timedelta(days=dt.weekday())  # 本周一
            week_idx = (dt - first_monday).days // 7
            rn = i + 2
            c_date = ws.cell(row=rn, column=1, value=dt)
            c_date.number_format = "yyyy-mm-dd"  # 真实日期格式，无时区偏移
            ws.cell(row=rn, column=2, value="周" + WEEKDAY_CN[dt.weekday()])
            ws.cell(row=rn, column=3, value=d.get("type", ""))
            ws.cell(row=rn, column=4, value=("/".join(d.get("projectNames") or [])) or "-")
            ws.cell(row=rn, column=5, value=d.get("manDays", 0))
            reason = d.get("manDayReason", "") or ""
            content = d.get("content", "") or ""
            c_reason = ws.cell(row=rn, column=6, value=reason)
            c_reason.alignment = _WRAP
            c_content = ws.cell(row=rn, column=7, value=content)
            c_content.alignment = _WRAP
            ws.cell(row=rn, column=8, value="是" if d.get("estimated") else "否")
            # 多行内容时按需加高行高
            lines = max(reason.count("\n") + 1, content.count("\n") + 1, 1)
            if lines > 1:
                ws.row_dimensions[rn].height = min(lines * 15 + 4, 120)
            if week_idx % 2 == 1:  # 按自然周(周一对齐)交替浅蓝底色
                for c in range(1, len(headers) + 1):
                    ws.cell(row=rn, column=c).fill = _BLUE
        _header_style(ws, len(headers), len(days))
        for c, w in enumerate([12, 8, 10, 24, 8, 24, 44, 10], start=1):
            ws.column_dimensions[get_column_letter(c)].width = w

        # 2) 总统计
        ws_sum = wb.create_sheet("总统计")
        ws_sum.append(["指标", "数值"])
        for row in [
            ("统计范围", f"{start} 至 {end}" if (start and end) else "-"),
            ("导出时间", datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("应出勤天数", total.get("requiredWorkDays", 0)),
            ("上班天数", total.get("workDays", 0)),
            ("请假天数", total.get("leaveDays", 0)),
            ("加班天数", total.get("overtimeDays", 0)),
            ("预估天数", total.get("estimatedDays", 0)),
            ("人天合计", total.get("manDays", 0)),
        ]:
            ws_sum.append(row)
        _header_style(ws_sum, 2, len(ws_sum["A"]) - 1)
        ws_sum.column_dimensions["A"].width = 16
        ws_sum.column_dimensions["B"].width = 28

        # 3) 按月统计
        ws_month = wb.create_sheet("按月统计")
        m_headers = ["月份", "上班", "请假", "加班", "预估", "人天", "应出勤"]
        ws_month.append(m_headers)
        for m in monthly:
            ws_month.append([m.get("month", ""), m.get("workDays", 0), m.get("leaveDays", 0),
                             m.get("overtimeDays", 0), m.get("estimatedDays", 0), m.get("manDays", 0),
                             m.get("requiredWorkDays", 0)])
        _header_style(ws_month, len(m_headers), len(monthly))
        for c, w in enumerate([14, 8, 8, 8, 8, 8, 8], start=1):
            ws_month.column_dimensions[get_column_letter(c)].width = w

        # 4) 按项目统计
        ws_proj = wb.create_sheet("按项目统计")
        ws_proj.append(["项目名称", "天数", "人天"])
        for p in by_project:
            ws_proj.append([p.get("projectName", ""), p.get("days", 0), p.get("manDays", 0)])
        _header_style(ws_proj, 3, len(by_project))
        ws_proj.column_dimensions["A"].width = 26
        ws_proj.column_dimensions["B"].width = 8
        ws_proj.column_dimensions["C"].width = 8

        # 5) 统计图表（原生 Excel 图表）
        ws_chart = wb.create_sheet("统计图表")
        ws_chart.append(m_headers)
        for m in monthly:
            ws_chart.append([m.get("month", ""), m.get("workDays", 0), m.get("leaveDays", 0),
                             m.get("overtimeDays", 0), m.get("estimatedDays", 0), m.get("manDays", 0)])
        month_last = len(monthly) + 1
        # 类型分布表（G/H 列，供饼图引用）
        ws_chart["G1"] = "类型"
        ws_chart["H1"] = "天数"
        dist = [
            ("上班", total.get("workDays", 0)),
            ("请假", total.get("leaveDays", 0)),
            ("加班", total.get("overtimeDays", 0)),
            ("预估", total.get("estimatedDays", 0)),
        ]
        for i, (label, val) in enumerate(dist, start=2):
            ws_chart.cell(row=i, column=7, value=label)
            ws_chart.cell(row=i, column=8, value=val)

        if monthly:
            bar = BarChart()
            bar.type = "col"
            bar.title = "按月出勤趋势"
            bar.y_axis.title = "天数"
            bar.x_axis.title = "月份"
            bar.add_data(Reference(ws_chart, min_col=2, max_col=5, min_row=1, max_row=month_last),
                         titles_from_data=True)
            bar.set_categories(Reference(ws_chart, min_col=1, min_row=2, max_row=month_last))
            bar.height = 8
            bar.width = 16
            ws_chart.add_chart(bar, "A" + str(month_last + 3))

        pie = PieChart()
        pie.title = "出勤类型分布"
        pie.add_data(Reference(ws_chart, min_col=8, min_row=1, max_row=5), titles_from_data=True)
        pie.set_categories(Reference(ws_chart, min_col=7, min_row=2, max_row=5))
        pie.height = 8
        pie.width = 12
        ws_chart.add_chart(pie, "G" + str(month_last + 3))

        for c in range(1, 9):
            ws_chart.column_dimensions[get_column_letter(c)].width = 12

        # 写出并返回
        buf = io.BytesIO()
        wb.save(buf)
        raw = buf.getvalue()
        filename = f"出勤明细_{start}_至_{end}.xlsx" if (start and end) else "出勤明细.xlsx"
        encoded = urllib.parse.quote(filename)
        return Response(
            content=raw,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded}",
                "Content-Length": str(len(raw)),
            },
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")
